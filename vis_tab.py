from tkinter import *
import time
from os import walk as wa, chdir as chd, getcwd as gtc, path as pas, makedirs as mkd
from re import match as mat
from openpyxl import Workbook as Wbk

cnt = ('00', '01', '02', '03', '04', '05', '06', '07', '08', '09', '10', '11',
       '12', '13', '14', '15', '16', '17', '18', '19', '20', '21', '22', '23')  # 北京时时间排列
cnt_utc = ('17', '18', '19', '20', '21', '22', '23', '00', '01', '02', '03', '04',
           '05', '06', '07', '08', '09', '10', '11', '12', '13', '14', '15', '16')  # UTC时间排列
month = ('Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
         'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec')  # 月份英语缩写按次序排列
clm = ('A', 'B', 'C', 'D', 'E', 'F', 'G', 'H')  # 小尾巴共有八列数据
prom_form = '输入VAISALA版本，新版输入"new"，旧版输入"old"（字符为小写字母，不要加引号），按回车键结束：'
prom_year = '输入需要统计能见度的年份（例如：2014），按回车键结束：'
prom_cont = '如果程序放在有"2012"、"2013"等文件夹的目录下，请直接按回车键，否则输入数据存放的目录：'
prom_over = '创建成功，请查看目录内对应文件，按回车键退出...'
rpt_name = 'REPORTS'  # 限定REPORT文件的文件名以什么开头，新版：AVIMET_REPORTS或者旧版：REPORTS（缺省）
rvr_name = 'RVR_RWY35'  # 限定RVR文件的文件名以什么开头，新版：RVR_METAR_35或者旧版：RVR_RWY35（缺省）
glo_path = gtc()  # 工作文件夹路径，缺省为当前程序所在路径
glo_form = 'old'  # 原始数据文件版本：new/old，缺省值为old
glo_year = '2014'  # 年份，缺省值为2014

#####################################################################################################################
"""
ver 1.x
以下部分为月总簿生成能见度数据汇总表格算法
"""


def year_dict(year):
    """
    ——创建一个新的年字典，数据结构为year--{month:{day:{hour:vis_data}}}，vis_data的初始值为NO DATA

    :param year: 需要统计能见度的年份，最好用字符串表示，如：'2014'
    :return: 一个三层嵌套的字典
    """
    year = str(year)
    y_dict = {}
    mth_dir_l = contents_names(year, 'dir')
    for i in mth_dir_l:
        day_file_l = find_list(contents_names(year + '/' + i, 'file'), rpt_name)
        m_dict = {}
        for j in day_file_l:
            d_dict = {}
            for k in cnt:
                d_dict[k] = 'NO DATA'
            m_dict[j[-6: -4]] = d_dict
        y_dict[i] = m_dict
    return y_dict


def data_to_dict(data: list,
                 mth_dct: dict):
    """
    ——将月字典中的日（键）-->小时（键）对应的值填充为对应的能见度值

    :param data: 以列表形式存储的一天的报数据
    :param mth_dct: 一个月的能见度数据字典，为年字典的（12个）值之一
    :return: 运行完成，返回1
    """
    for x in data:
        hour = ''
        for y in x:
            if find_sp(y):
                break
            elif find_time(y):
                hour = y
            elif find_vis(y) and hour:
                mth_dct[hour[0: 2]][hour[2: 4]] = get_vis(y)
                break
            else:
                continue
    return 1


def fill_month_dict(year,
                    mth: str,
                    y_dct: dict):
    """
    ——根据给定的参数（年，月）填充年字典y_dct中的相应月份的所有值

    :param year: 需要统计能见度的年份，最好用字符串表示，如：'2014'
    :param mth: 当前统计的月份，为英语缩写的字符串，应与对应文件夹名称完全一致
    :param y_dct: 三层嵌套的字典，数据结构为year--{month:{day:{hour:vis_data}}}
    :return: 如函数正常运行，返回整数1
    """
    year = str(year)
    rpt_list = find_list(contents_names(year + '/' + mth, 'file'), rpt_name)
    for i in rpt_list:
        data_to_dict(file_to_list(year + '/' + mth + '/' + i), y_dct[mth])
    return 1


def fill_year_dict(year,
                   y_dct: dict):
    """
    ——根据给定的参数（年）填充年字典y_dct

    :param year: 需要统计能见度的年份，最好用字符串表示，如：'2014'
    :param y_dct: 三层嵌套的字典，数据结构为year--{month:{day:{hour:vis_data}}}
    :return: 返回填充好的年字典
    """
    year = str(year)
    for mth in contents_names(year, 'dir'):
        fill_month_dict(year, mth, y_dct)
    return y_dct


def contents_names(f_path: str,
                   k: str):
    """
    ——对于指定路径f_path,根据k的值，以列表形式返回所有文件名/目录名/路径，本场景k=dir/files/root

    :param f_path: 指定的文件路径
    :param k: 需要取出名字的项目，dir=所有文件夹名、files=所有文件名、root=路径名称
    :return: 将所有取出的名字以一个列表的形式返回
    """
    for r, d, f in wa(f_path):
        if 'dir' in k:
            return list(d)
        elif 'fil' in k:
            return list(f)
        elif 'root' in k:
            return list(r)
        else:
            print('输入错误！')


def find_list(cnt_names: list,
              tgt_name: str):
    """
    ——从所给列表元素中找出以指定字符串开头的元素，将这些字符串组合成一个新的列表

    :param cnt_names: 原始列表
    :param tgt_name: 指定的的字符串
    :return: 符合条件的元素组成的列表
    """
    tgt_list = []
    for n in cnt_names:
        if n.startswith(tgt_name):
            tgt_list += [n]
    return tgt_list


def file_to_list(file: str):
    """
    ——将文件中的内容分解为一个二维列表，第一层为每行数据，第二层为每行的每个单词

    :param file: 存储每日电码格式报文的数据文件，如：REPORT01.his
    :return: 返回列表嵌套，外层每个元素为上述文件中的一行（每行为一个内层列表），内层每个元素为每行的每个单词
    """
    data_line = []
    with open(file, 'r') as f:
        for line in f.readlines():
            data_line += [line.split()]
    return data_line


def find_sp(s: str):
    """
    ——判断是否为SPECI标识符

    :param s: 报文单词
    :return: 如果是SP，返回真，否则返回假
    """
    if s == 'SPECI':
        return True
    else:
        return False


def find_time(s: str):
    """
    ——判断是否为时间组

    :param s: 报文单词
    :return: 如果是时间组，返回真，否则返回假
    """
    if mat('\\d\\d\\d\\d\\d\\dZ', s) and len(s) == 7:
        return True
    else:
        return False


def find_vis(s: str):
    """
    ——判断是否为能见度组

    :param s: 报文单词
    :return: 如果是能见度组，返回真，否则返回假
    """
    if s == 'CAVOK' or s == '9999' or s == '10000':
        return True
    elif mat('\\d\\d\\d\\d', s) and len(s) == 4:
        return True
    else:
        return False


def get_vis(s: str):
    """
    ——给定表示能见度的单词，推算出能见度值

    :param s: 可以表示能见度值的字符串，包括四位能见度电码取值、CAVOK、9999等
    :return: 表示能见度值的字符串
    """
    if s == 'CAVOK' or s == '9999' or s == '10000':
        return '10000'
    else:
        return s


def print_ln(x: any):
    """
    ——打印后换行，如果是字典，打印其所有的值

    :param x: 任意变量
    :return: 运行完成返回1
    """
    if isinstance(x, dict):
        for i in x.keys():
            print(x[i])
    else:
        for i in x:
            print(i)
    print('\n')
    return 1


def year_workbook():
    """
    ——创建一个工作簿book，在其中依次创建12个月工作表，将所有工作表存入一个字典ws_dct——{月份:月工作表}

    :return: 工作簿，包含12个工作表的字典
    """
    book = Wbk()
    ws_dct = {}
    for i in month:
        ws_dct[i] = book.create_sheet(i, -1)
    return book, ws_dct


def utc_to_cst(day, hour):
    """
    ——将UTC（世界协调时）转化为CST（北京时）

    :param day: UTC日期
    :param hour: UTC小时
    :return: CST日期，CST小时
    """
    day = int(day)
    hour = int(hour)
    if hour < 17:
        hour += 8
    else:
        hour -= 16
        day += 1
    return day, hour


def vis_to_int(dt: str):
    """
    ——如果VIS值不为NO DATA，将值转化为数字

    :param dt: 原始值
    :return: VIS值的数字/NO DATA
    """
    data = int(dt) if dt != 'NO DATA' else dt
    return data


def save_to_excel(year):
    """
    ——将一年的能见度数值按照既定的数据结构，保存到Excel中

    :param year: 年份
    :return: 储存了年数据的字典
    """
    year = str(year)
    dct_y = year_dict(year)  # 创建年字典
    fill_year_dict(year, dct_y)  # 将数据填入年字典
    wb_y, ws_dct_y = year_workbook()  # 创建工作簿，创建12个工作表
    for i in dct_y.keys():
        for ro in dct_y[i].keys():
            for co in dct_y[i][ro].keys():
                day, hour = utc_to_cst(ro, co)  # 此处完成世界/北京时转换，使世界时的能见度值对应到北京时的行列
                ws_dct_y[i].cell(row=int(day),  # 北京时日期对应行
                                 column=int(hour) + 1,  # 北京时小时对应列
                                 value=vis_to_int(dct_y[i][ro][co]))  # 填入数据，数据对应的时间是UTC时间
            ws_dct_y[i].cell(row=int(ro),
                             column=1,
                             value=ro + '日')  # 在第一列每一行开头填入日期数据
    wb_y = replace_tail(wb_y)  # 将小尾巴移动到下一月开头
    for i in dct_y.keys():
        ws_dct_y[i].insert_rows(1)  # 在第一行上方插入一行
        for co in range(25):
            if co != 0:
                ws_dct_y[i].cell(row=1,
                                 column=co + 1,
                                 value=cnt_utc[co - 1])  # 分别填入UTC小时时次
            else:
                ws_dct_y[i].cell(row=1,
                                 column=co + 1,
                                 value='日期\\时次')  # 在第一行第一列填入表头
    make_dir('result/vis/')
    wb_y.save('result/vis/' + year + '年能见度汇总(CTS).xlsx')  # 保存为一个Excel，位置在原始文件夹/result/vis/
    return dct_y


def replace_tail(wb):
    """
    ——将每月最后几个数值替换到下一个月的开头，最后一个月的单独存一个工作表

    :param wb: 替换前的workbook
    :return: 替换后的workbook
    """
    for wsn in month:
        wmr = str(wb[wsn].max_row)
        wb[wsn]['A' + wmr].value = '01日'  # 为最后一行数据列首添加"01日"
        for cl in clm:
            if wsn == 'Dec':
                wb['Sheet'][cl + '1'].value = wb[wsn][cl + wmr].value  # 十二月的小尾巴复制到sheet的第一行
            else:
                wsn_next = month[month.index(wsn) + 1]  # 得出wsn下一个月的字母缩写
                wb[wsn_next][cl + '1'].value = wb[wsn][cl + wmr].value  # 当月的小尾巴复制到下一个月的开头
        wb[wsn].delete_rows(int(wmr))  # 删除小尾巴
    return wb


def year_data_to_excel(path: str,
                       year: str):
    """
    ——完成外部输入数据的初始化，如外部数据空缺，则启用缺省值

    ——将给定的目录下的给定年份的能见度数据统计到Excel中

    :param path: 工作目录，即存放各个年份原始数据的目录
    :param year: 需要生成的年份
    :return: 年份数据字典，年份
    """
    year = '2014' if year == '' else year  # 如果没有指定年份，默认采用2014
    chd(gtc() if path == '' else path)  # 如果没有指定目录，则读取当前程序所在目录
    y_dct = save_to_excel(year)  # 运行主程序算法
    return y_dct, year


def tab_it():  # 图形界面下，该函数不运行，命令行界面下，运行该函数以进行运算
    """
    ——采用命令行执行时（需要将图形界面屏蔽），使用输入值作为参数

    ——当监测到某个单元格是NO DATA时，将该单元格代表的时次显示出来，保存到一个txt文本中

    :return: 正确执行返回1
    """
    global rpt_name
    rpt_name = 'AVIMET_REPORTS' if input(prom_form) == 'new' else 'REPORTS'
    y_dct_p, year_p = year_data_to_excel(input(prom_cont), input(prom_year))
    for mth_p in y_dct_p.keys():
        for day_p in y_dct_p[mth_p].keys():
            for hour_p in y_dct_p[mth_p][day_p].keys():
                state = mth_p + ' ' + day_p + '' + hour_p + 'UTC is NO DATA'  # 需要显示的句子statement
                txt_file = open(year_p + 'NA.txt', 'a')
                print(state, file=txt_file) if y_dct_p[mth_p][day_p][hour_p] == 'NO DATA' else 0
    input(prom_over)
    return 1


def vis_tab_it():
    """
    ——配合图形界面执行，输入参数采用全局变量传入图形界面输入的参数

    ——当监测到某个单元格是NO DATA时，将该单元格代表的时次显示在图形界面

    :return:
    """
    global y_d  # , rpt_name
    # rpt_name = 'AVIMET_REPORTS' if glo_form == 'new' else 'REPORTS'
    # 已在初始化功能中定义该值，无需再次定义，相关函数：vis_tab_init
    y_dct_p, year_p = year_data_to_excel(glo_path, glo_year)
    delete_text_2()
    update_txt(text_2,
               glo_year + ':')
    for mth_p in y_dct_p.keys():
        for day_p in y_dct_p[mth_p].keys():
            for hour_p in y_dct_p[mth_p][day_p].keys():
                state = mth_p + ' ' + day_p + '' + hour_p + 'UTC is NO DATA'  # 需要显示的句子statement
                update_txt(text_2, state) if y_dct_p[mth_p][day_p][hour_p] == 'NO DATA' else 0
    update_txt(text_2,
               '')
    update_txt(text_3,
               '能见度表格创建成功，请查看目录内对应文件!')
    y_d = y_dct_p
    return 1


#################################################################################################################
"""
ver 2.1
以下部分为图形界面所用到的函数
"""

y_list = []
file_path = ''
y_d = {}
m_d = d_d = h_d = ''
data_len = 0
this_line = 1


def say_hi():
    """
    ——测试
    """
    print('目录：', entry1_2.get())
    print('功能：', v1_3.get())
    print('数据格式：', v1_5.get())
    print('年份：', v1_4.get())


def find_year_list(f_path: str):
    """
    ——从文件夹中发现已存在的年份（ver2.2新增 筛选出年份文件夹：符合文件夹名称由4个数字组成且名称总长度为4的文件夹）

    :param f_path: 图形界面输入的工作目录值
    :return: 年份名称组成的列表
    """
    list_1 = contents_names(f_path if f_path != '' else gtc(), 'dir')
    year_list = []
    for i in list_1:
        year_list += [i] if mat('\\d\\d\\d\\d', i) and len(i) == 4 else []
    return year_list


def vis_tab_init(path_no: str,
                 form_no: int,
                 year_no: int):
    """
    ——初始化输入值，将图形界面的输入值转换并赋值给全局变量

    :param path_no: 路径项的图形界面输入值（str）
    :param form_no: 数据格式按钮选择项值（int）
    :param year_no: 年份按钮选择项的值（int）
    :return:
    """
    # ver2.2更新，初始化同时定义全局函数
    # 限定REPORT文件的文件名以什么开头，新版：AVIMET_REPORTS或者旧版：REPORTS（缺省）
    # 限定RVR文件的文件名以什么开头，新版：RVR_METAR_35或者旧版：RVR_RWY35（缺省）
    global glo_path, glo_form, glo_year, rpt_name, rvr_name
    if path_no != '':
        glo_path = path_no
    if form_no == 1:
        glo_form = 'new'
        rpt_name = 'AVIMET_REPORTS'
        rvr_name = 'RVR_METAR_35'
    elif form_no == 2:
        glo_form = 'old'
        rpt_name = 'REPORTS'
        rvr_name = 'RVR_RWY35'
    glo_year = y_list[year_no]
    return 1


def update_txt(text: Text,
               texts: str):
    """
    ——在一个text组件中末尾更新一段texts，更新完成后锁定text

    :param text: tkinter中的类（class）Text，该组件的一个实例
    :param texts: 需要显示在组件最后一行的文本内容
    :return: 正常运行返回1
    """
    text.config(state=NORMAL)
    text.insert('end', texts + '\n')
    text.config(state=DISABLED)
    text.see('end')
    return 1


def mk_dir():  # 提交目录
    """
    ——点击提交按钮后，显示文本，提示目录当前设置值
    """
    global file_path
    file_path = entry1_2.get()
    if file_path == '':
        file_path = gtc()
    update_txt(text_3,
               '工作目录已设置为：\''
               + file_path + '\'\n'
               + '请确认该目录下存放了原始数据年份文件夹！')
    return 1


def para_init():  # 初始化
    """
    ——点击初始化按钮后，对各项全局参数进行初始化赋值

    ——显示文本，提示各个参数当前设置值
    """
    vis_tab_init(file_path,
                 v1_5.get(),
                 v1_4.get())
    update_txt(text_3,
               '初始化完成：'
               + '\n工作目录：\'' + glo_path + '\''
               + '\n数据格式：\'' + glo_form + '\''
               + '\n年份：\'' + glo_year + '\'')
    return 1


def old_form():  # 旧格式
    """
    ——点击旧格式按钮后，显示文本
    """
    update_txt(text_3,
               '已设定年份原始数据为旧数据格式(2012年2月~2017年6月)，请核实！')
    return 'old'


def new_form():  # 新格式
    """
    ——点击新格式按钮后，显示文本
    """
    update_txt(text_3,
               '已设定年份原始数据为新数据格式(2017年7月至今)，请核实！')
    return 'new'


def func_1():  # 校对
    """
    ——点击校对按钮后，显示文本
    """
    text_2.place(relx=0.01,
                 rely=0.07,
                 relheight=0.8,
                 relwidth=0.49)
    frame2_1.place(relx=0.51,
                   rely=0.07,
                   relheight=0.8,
                   relwidth=0.48)
    update_txt(text_3,
               '软件执行后可继续对"NO DATA"的数据进行手动校对，操作方法如下：\n'
               '1.点按\'开始校对\'按钮\n'
               '2.根据显示的时次信息，查找相应的数据，填入文本框\n'
               '3.点按\'更新能见度\'按钮\n'
               '4.按\'下一个\'校对下一个数据\n'
               '5.按\'上一个\'检查前面的数据\n'
               '6.全部校对完成后，按\'保存校对后的Excel\'保存数据\n'
               '7.校对完成并存为Excel前，请不要清屏，否则请重新生成！')
    return 1


def func_2():  # 不校对
    """
    ——点击不校对按钮后，显示文本
    """
    text_2.place(relx=0.01,
                 rely=0.07,
                 relheight=0.8,
                 relwidth=0.98)
    frame2_1.lower()
    update_txt(text_3,
               '软件执行后仅显示"NO DATA"的数据，请在生成的Excel文件中校对')
    return 2


def mk_year():  # 年份
    """
    ——点击年份按钮后，显示文本
    """
    update_txt(text_3,
               '已重新选择年份，请关注数据格式是否有变化！')
    return 2


def search_year(frame: LabelFrame,
                v: IntVar):  # 加载可用年份
    """
    ——点击加载可用年份按钮后，根据输入的目录查找该目录下年份文件夹的名称，显示可用的年份

    :param frame: 年份选择按钮显示在的框架
    :param v: 初始显示选中的年份
    :return: 运行正确返回1
    """
    global y_list
    y_list = find_year_list(file_path)
    try:
        y_list.sort()
        rb = {}
        for widget in frame.winfo_children():
            widget.destroy()  # 清空框架中原本的子内容
        for i in range(len(y_list)):
            rb[i] = Radiobutton(frame,
                                text=y_list[i],
                                fg='#111111',
                                variable=v,
                                value=i,
                                command=mk_year)
            if i % 2 == 0:
                rb[i].place(relx=0.10,
                            rely=0.07 + 0.1 * i / 2)
            else:
                rb[i].place(relx=0.60,
                            rely=0.07 + 0.1 * (i - 1) / 2)
        update_txt(text_3,
                   '可用年份已全部加载，请选择一个年份开始')
    except AttributeError:
        update_txt(text_3,
                   '该目录下没有年份文件夹，请重新输入目录，或留空！输入后请点按\'提交\'按钮。')
    return 1


def get_time():
    """
    ——以HHMMSS DDMMYY的格式更新时间，每1000毫秒更新一次

    :return: 空
    """
    var_time.set(time.strftime('%H:%M:%S %d/%m/%y'))  # 获取当前时间
    root.after(1000, get_time)  # 每隔1s调用函数自身获取时间


def goto_before():  # 上一个
    """
    ——加载列表中的上一个数据

    :return: 1
    """
    global data_len, this_line, y_d, m_d, d_d, h_d
    if this_line <= data_len:
        if this_line > 2:
            this_line -= 1
        else:
            this_line = data_len
    else:
        this_line = 1
    if this_line == 1:
        change_l2_1('////', '////')
        this_line += 1
        update_txt(text_3,
                   '未找到可校对的数据，请重试，请按提示的步骤执行！')
    elif this_line <= data_len:
        mth_0 = float(this_line)
        mth_1 = this_line + 0.3
        day_0 = this_line + 0.4
        day_1 = hour_0 = this_line + 0.6
        hour_1 = this_line + 0.8
        # data_0 = this_line + 0.15
        # data_1 = this_line + 0.22
        m_d = get_t2(mth_0, mth_1)
        d_d = get_t2(day_0, day_1)
        h_d = get_t2(hour_0, hour_1)
        # data = get_t2(data_0, data_1)
        change_l2_1(glo_year + '-' + m_d + '\n' + d_d + h_d + '00Z', y_d[m_d][d_d][h_d])
    else:
        change_l2_1('error', 'error')
        update_txt(text_3,
                   '数据出错，请重试！')
    return 1


def update_vis():  # 更新能见度
    """
    ——将能见度值更新到字典中

    :return: 1
    """
    global y_d, m_d, d_d, h_d
    vis_d = entry2_1_1.get()
    try:
        int(vis_d)
        y_d[m_d][d_d][h_d] = vis_d
        update_txt(text_3,
                   glo_year + '-' + m_d + '-' + d_d + h_d + '00Z的能见度已更新为：' + str(int(y_d[m_d][d_d][h_d])))
    except ValueError:
        update_txt(text_3,
                   '请输入正确的能见度值！')
    return 1


def goto_next():  # 下一个
    """
    ——加载列表中的下一行数据

    :return: 1
    """
    global data_len, this_line, y_d, m_d, d_d, h_d
    if this_line < data_len:
        this_line += 1
    elif this_line == data_len:
        this_line = 2
    else:
        this_line = 1
    if this_line == 1:
        change_l2_1('////', '////')
        this_line += 1
        update_txt(text_3,
                   '未找到可校对的数据，请重试，请按提示的步骤执行！')
    elif this_line <= data_len:
        mth_0 = float(this_line)
        mth_1 = this_line + 0.3
        day_0 = this_line + 0.4
        day_1 = hour_0 = this_line + 0.6
        hour_1 = this_line + 0.8
        # data_0 = this_line + 0.15
        # data_1 = this_line + 0.22
        m_d = get_t2(mth_0, mth_1)
        d_d = get_t2(day_0, day_1)
        h_d = get_t2(hour_0, hour_1)
        # data = get_t2(data_0, data_1)
        change_l2_1(glo_year + '-' + m_d + '\n' + d_d + h_d + '00Z', y_d[m_d][d_d][h_d])
    else:
        change_l2_1('error', 'error')
        update_txt(text_3,
                   '数据出错，请重试！')
    return 1


def save_txt():  # 存文本
    """
    ——将校对文本框中的显示内容存储到一个TXT文件中

    :return: 1
    """
    print(text_2.get(1.0, 'end'), file=open(glo_year + ' NO DATA.txt', 'w'))
    return 1


def delete_text_2():  # 清屏
    """
    ——删除校对文本框文档内容

    :return: 1
    """
    text_2.config(state=NORMAL)
    text_2.delete(1.0, 'end')
    text_2.config(state=DISABLED)
    return 1


def start_update_vis():  # 开始校对
    """
    ——记录年份，统计无数据记录的时次信息，从年数据字典中读取相应第一条时次的数据并显示

    :return: 1
    """
    global data_len
    # print(text_2.index('end'))
    # print(text_2.index('end').__class__)
    # print(text_2.get(float(text_2.index('end')) - 3, float(text_2.index('end'))) + 'end')
    data_len = int(float(text_2.index('end')) - 3)  # 最后一条数据的行数
    if data_len > 1:
        change_l2_1('点下一个',
                    '')
        update_txt(text_3,
                   '请点按\'下一个\'开始校对能见度')
    else:
        change_l2_1('////',
                    '////')
        update_txt(text_3,
                   '没有可校对的数据，请按照提示的步骤操作！')
    return 1


def save_change_excel():  # 保存校对后的Excel
    """
    ——保存校对后的Excel，如果未进行逐个校对，则数据不会有改变

    :return: 1
    """
    wb_y, ws_dct_y = year_workbook()  # 创建工作簿，创建12个工作表
    for i in y_d.keys():
        for ro in y_d[i].keys():
            for co in y_d[i][ro].keys():
                day, hour = utc_to_cst(ro, co)  # 此处完成世界/北京时转换，使世界时的能见度值对应到北京时的行列
                ws_dct_y[i].cell(row=int(day),  # 北京时日期对应行
                                 column=int(hour) + 1,  # 北京时小时对应列
                                 value=vis_to_int(y_d[i][ro][co]))  # 填入数据，数据对应的时间是UTC时间
            ws_dct_y[i].cell(row=int(ro),
                             column=1,
                             value=ro + '日')  # 在第一列每一行开头填入日期数据
    wb_y = replace_tail(wb_y)  # 将小尾巴移动到下一月开头
    for i in y_d.keys():
        ws_dct_y[i].insert_rows(1)  # 在第一行上方插入一行
        for co in range(25):
            if co != 0:
                ws_dct_y[i].cell(row=1,
                                 column=co + 1,
                                 value=cnt_utc[co - 1])  # 分别填入UTC小时时次
            else:
                ws_dct_y[i].cell(row=1,
                                 column=co + 1,
                                 value='日期\\时次')  # 在第一行第一列填入表头
    make_dir('result/vis/')
    wb_y.save('result/vis/' + glo_year + '年能见度汇总(CTS)-已校对.xlsx')  # 保存为一个Excel，位置在原始文件夹/result/vis/
    return 1


def get_t2(x: float,
           y: float):
    """
    ——从text_2中取从x到y的对应数据

    :param x: 开头
    :param y: 结尾
    :return: 数据段
    """
    return text_2.get(index1=x,
                      index2=y)


def change_l2_1(s1: str,
                s2: str):
    """
    ——改变校对栏两个label的显示内容

    :param s1: label_2_1_5的显示内容
    :param s2: label_2_1_6的显示内容
    :return: 1
    """
    v_2_1_5.set(s1)
    v_2_1_6.set(s2)
    return 1


#################################################################################################################
"""
ver 2.2
新增功能：读取所有整点RVR值，存为Excel
ver 2.2.1
新增功能：查找RVR异常值，存为txt
"""


def make_dir(path: str):
    """
    ——判断是否存在文件夹如果不存在则创建为文件夹

    :param path: 文件夹路径，如果是相对路径上级为程序工作目录
    :return: 文件夹路径
    """
    folder = pas.exists(path)
    if not folder:
        mkd(path)  # makedirs 创建文件时如果路径不存在会创建这个路径
    else:
        pass
    return path


def rvr_file_to_dict(file: str):
    """
    ——从文件中筛选出世界时整点（00～23时）数据，分门别类存为一个字典，具体如下：

    ——将文件中的内容分解为一个二维列表，第一层为每行数据列表组成的列表，第二层为每行的每个元素，以制表符为分割

    ——数据列表结构中第一行数据忽略，从第二行开始按照规则筛选数据存入字典

    ——建立一个字典嵌套，第一层键为时间，第二层键为列表第一行每个元素（标题），第三层为相应时间和标题下的值

    ——cnt中的整点值对应的数据行（列表元素行）为：cnt * 60 + 2

    ——新增：以10min为时距，查找RVR小于50的数据，存到result/rvr_lost/文件夹下的文件中

    :param file: 存储vaisala原始数据的文件名，如：RVR_RWY35_01.his
    :return: 返回字典嵌套:{时间:{标题:值}}
    """
    data_line = []
    day_rvr_dict = {}
    make_dir('result/rvr_lost/')
    with open(file, 'r') as f:
        for line in f.readlines():
            data_line += [line.split('\t', 9)]  # 9为RVR文件前9段数据有意义，最后一段为剩余所有数据，根据不同的文件进行更改
    count = 0
    for i in range(len(data_line)):
        if data_line[i][0].endswith('0:00'):
            try:
                rvr10 = data_line[i][7]
                rvr1 = data_line[i][4]
                if rvr10 == '0' or rvr10 == '25' or rvr10 == ' ' or rvr10 == '' or rvr10 is None:
                    print(data_line[i][0] + 'RVR10A: |' + data_line[i][7] + '|',
                          file=open('result/rvr_lost/' + glo_year + ' M0050 10min 10a.txt', 'a'))
                if rvr1 == '0' or rvr1 == '25' or rvr1 == ' ' or rvr1 == '' or rvr1 is None:
                    print(data_line[i][0] + 'RVR1A: |' + data_line[i][4] + '|',
                          file=open('result/rvr_lost/' + glo_year + ' M0050 10min 1a.txt', 'a'))
            except IndexError:
                pass
            except ValueError:
                pass
        if data_line[i][0].endswith(':00:00'):
            count += 1
            hour = data_line[i][0][11:13]  # i行判定为整点，i行第一个字符串为世界时时间，第12到13位为小时字符
            day_rvr_dict[hour] = {}
            for j in range(10):
                title = data_line[1][j]  # 第2行为标题行
                day_rvr_dict[hour][title] = data_line[i][j]
    if count < 24:
        update_txt(text_2,
                   data_line[-1][0][0:11] + 'RVR数据有缺失，请复查！')
        for i in cnt:
            if day_rvr_dict.get(i) is None:
                update_txt(text_2,
                           'hour:' + i + ' 无数据！已补充为\'NO DATA\'，请复查！')
                day_rvr_dict[i] = {}
                for j in range(10):
                    title = data_line[1][j]  # 第2行为标题行
                    day_rvr_dict[i][title] = 'NO DATA'
    for i in cnt:
        r = day_rvr_dict[i]
        if r.get('RVR_10A') == ' ':
            update_txt(text_2,
                       r.get('CREATEDATE') + '  监测到RVR_10A为空，请复查！')

    """
    for i in cnt:
        day_rvr_dict[i] = {}
        for j in range(10):
            day_rvr_dict[i][data_line[1][j]] = data_line[int(i) * 60 + 2][j]
            # 适用于新格式的数据，旧格式不需要*4，明天改
            # 数据不能缺失，缺失会报错，明天改
            # 思路1：改为直接根据时间后缀为:00:00
            # 思路2：根据抛出错误，可以判定是否缺失数据
            # 思路3：每次处理完一天的数据，检验是否为24个数据，不是，则用text3描述抛出错误
            # 改变字符的类型，是否将RVR处理成整数，或者处理成M0050/P2000N等格式
    """
    return day_rvr_dict


def month_rvr_to_dict(year: str,
                      mth: str):
    """
    ——根据给定的月份，制作一个字典：

    ——从月份文件夹中找到符合条件的文件（28～31个RVR数据文件），分别将每个文件中符合条件的数据制作一个字典

    ——将这些字典合并成一个月字典：{日期:{时间:{标题:值}}}

    :param year: 需要处理的年份名称（文件夹名称）
    :param mth: 需要处理的月份名称（文件夹名称）
    :return: 返回字典嵌套：{日期:{时间:{标题:值}}}
    """
    mth_rvr_dict = {}
    try:
        day_file_list = find_list(contents_names(year + '/' + mth, 'file'), rvr_name)
        try:
            # day_file_list.sort()  # 以下算法可以不用排序，如果用索引访问list，则最好先排序
            for i in day_file_list:
                mth_rvr_dict[i[-6:-4]] = rvr_file_to_dict(year + '/' + mth + '/' + i)
                # e.g. i: RVR_RWY35_01.his, RVR_METAR_35_01.his,i为文件名，其[-6:-4]字符为日期，日期为键，该日的字典为值
        except AttributeError:  # 应对list中没有任何元素的情况，说明月文件夹下为空或不包含指定的文件，抛出该错误
            update_txt(text_3,
                       '未找到可用文件，请检查月份文件夹下是否存在相关文件。')
    except TypeError:
        update_txt(text_3,
                   year + ' ' + mth + ' ：该月份文件夹不存在，请复查！')
    return mth_rvr_dict


def year_rvr_to_dict(year: str):
    """
    ——根据给定的年份，制作一个字典：{月份:{日期:{时间:{标题:值}}}}

    :param year: 需要处理的年份名称（文件夹名称）
    :return: 返回字典嵌套：{月份:{日期:{时间:{标题:值}}}}
    """
    year_rvr_dict = {}
    # mth_file_list = find_list(contents_names(year + '/' + mth, 'file'), rvr_name)
    try:
        for m in month:
            year_rvr_dict[m] = month_rvr_to_dict(year, m)
    except AttributeError:
        update_txt(text_3,
                   '未找到可用文件，请检查该年文件夹下，月份文件夹是否存在。')
    return year_rvr_dict


def make_rvr_book():
    """
    ——将字典（{月份:{日期:{时间:{标题:值}}}}）中的按照格式内容写入到一个表格中

    :return: 返回1
    """
    year = glo_year
    dct = year_rvr_to_dict(year)
    book = Wbk()
    # sheet = book.create_sheet(year, 0)
    sheet = book['Sheet']
    count = 1
    sheet.cell(row=count,
               column=1,
               value='CREATEDATE(UTC)')
    sheet.cell(row=count,
               column=2,
               value='RVR_10A')
    for i in range(len(month)):
        day_list = []
        for d in dct[month[i]].keys():
            day_list += [d]
        try:
            day_list.sort()
        except AttributeError:
            update_txt(text_3,
                       year + '年' + month[i] + '月目录下无数据！')
        for j in range(len(day_list)):
            for k in range(24):
                count += 1
                sheet.cell(row=count,
                           column=1,
                           value=dct[month[i]][day_list[j]][cnt[k]]['CREATEDATE'])
                sheet.cell(row=count,
                           column=2,
                           value=dct[month[i]][day_list[j]][cnt[k]]['RVR_10A'])
    make_dir('result/rvr/')
    book.save('result/rvr/' + year + 'RVR统计结果.xlsx')  # 保存为一个Excel，位置在原始文件夹/result/rvr/
    update_txt(text_3,
               'RVR年数据表格创建成功，请在原始数据目录查看相关文件！\n'
               '注意：RVR数据未设置校对功能，一般为原始数据缺失，无法校对，请手动复查！')
    # print(type(dct[month[5]]['12']['07'].get('RVR_10A')), dct[month[5]]['12']['07'].get('RVR_10A'))
    return book


#################################################################################################################
"""
ver 2.1
以下部分为图形界面
"""

root = Tk()
root.title('VIS_TAB')
root.geometry('1280x720')  # 窗体尺寸，中间为x，不是*

label_time = Label(root)
label_time.place(relx=0.469,
                 rely=-0.003)
var_time = StringVar()
label_time.config(textvariable=var_time,
                  fg='#222222',
                  font=('Arial', 9))
get_time()

frame3 = LabelFrame(root,
                    relief=SOLID)  # 框架
frame3.place(relx=0.31,
             rely=0.71,
             relheight=0.27,
             relwidth=0.68)

label3_1 = Label(frame3,
                 text='信息提示',
                 bg='#eeeeee',
                 fg='#111111',
                 justify=LEFT,
                 relief=FLAT)
label3_1.place(relx=0.01,
               rely=0.02,
               relwidth=0.98)

text_3 = Text(frame3,
              state=DISABLED)
text_3.place(relx=0.01,
             rely=0.15,
             relheight=0.80,
             relwidth=0.98)
update_txt(text_3,
           '请先阅读本说明文本，再进行数据操作！！！' * 3 + '重要的事情说三遍！\n'
           '数据路径请填写包含年份文件夹的目录的完整路径，如果软件也位于该文件夹，则可以留空不填！\n'
           '新生成的文件保存在该文件夹，无提示覆盖旧的同名文件，请做好文件备份，如覆盖，重新生成即可！\n'
           '确定路径后请点按\'加载可用年份\'，如果显示出年份选项，则表示路径正确。否则请检查文件路径。\n'
           '选择年份后请注意该年份对应的数据格式（尤其是2017），点按\'新\'或\'旧\'可查看具体对应时间。\n'
           '特别注意2017年，选择\'旧\'、\'新\'后会分别生成前、后半年数据，生成后需要另存，再做另一半。\n'
           '选择好年份、数据格式后，点按\'初始化\'，查看显示的信息是否正确，点按\'生成\'即可生成表格。\n'
           '如果选择了\'校对\'，则在生成了Excel后，可继续对NO DATA数据进行逐个校对，请按提示完成校对。\n'
           '在完成了\'初始化\'操作之后（无需点按\'生成\'），通过点按\'RVR\'按钮，生成当年的RVR数据。\n')
sb_3 = Scrollbar(text_3,
                 command=text_3.yview)
sb_3.pack(side='right',
          fill='y')
text_3.config(yscrollcommand=sb_3.set)

frame1 = LabelFrame(root,
                    relief=SOLID)  # 框架
frame1.place(relx=0.01,
             rely=0.02,
             relheight=0.96,
             relwidth=0.29)

label1_1 = Label(frame1,
                 text='参数设置',
                 bg='#eeeeee',
                 fg='#111111',
                 justify=CENTER,
                 relief=FLAT)
label1_1.place(relx=0.01,
               rely=0.02,
               relwidth=0.98)

label1_2 = Label(frame1,
                 text='数据路径',
                 bg='#eeeeee',
                 fg='#111111',
                 justify=LEFT,
                 relief=FLAT)
label1_2.place(relx=0.05,
               rely=0.12)

entry1_2 = Entry(frame1,
                 width=16)
entry1_2.place(relx=0.3,
               rely=0.12)

button1_2 = Button(frame1,
                   text='提交',
                   fg='#111111',
                   command=mk_dir)
button1_2.place(relx=0.8,
                rely=0.12)

label1_3 = Label(frame1,
                 text='功能选择',
                 bg='#eeeeee',
                 fg='#111111',
                 justify=LEFT,
                 relief=FLAT)
label1_3.place(relx=0.05,
               rely=0.22)

v1_3 = IntVar()
v1_3.set(2)

rb1_3_1 = Radiobutton(frame1,
                      text='校对',
                      fg='#111111',
                      variable=v1_3,
                      value=1,
                      command=func_1)
rb1_3_1.place(relx=0.40,
              rely=0.22)

rb1_3_2 = Radiobutton(frame1,
                      text='不校对',
                      fg='#111111',
                      variable=v1_3,
                      value=2,
                      command=func_2)
rb1_3_2.place(relx=0.40,
              rely=0.27)

label1_4 = Label(frame1,
                 text='数据年份',
                 bg='#eeeeee',
                 fg='#111111',
                 justify=LEFT,
                 relief=FLAT)
label1_4.place(relx=0.05,
               rely=0.32)

v1_4 = IntVar()
v1_4.set(0)

frame1_4 = LabelFrame(frame1,
                      relief=GROOVE)
frame1_4.place(relx=0.25,
               rely=0.39,
               relheight=0.41,
               relwidth=0.67)

button1_4 = Button(frame1,
                   text='加载可用年份',
                   fg='#111111',
                   command=lambda: search_year(frame1_4, v1_4))
button1_4.place(relx=0.3,
                rely=0.32)

label1_5 = Label(frame1,
                 text='数据格式',
                 bg='#eeeeee',
                 fg='#111111',
                 justify=LEFT,
                 relief=FLAT)
label1_5.place(relx=0.05,
               rely=0.82)

v1_5 = IntVar()
v1_5.set(2)

rb1_5_1 = Radiobutton(frame1,
                      text='新',
                      fg='#111111',
                      variable=v1_5,
                      value=1,
                      command=new_form)
rb1_5_1.place(relx=0.40,
              rely=0.82)

rb1_5_2 = Radiobutton(frame1,
                      text='旧',
                      fg='#111111',
                      variable=v1_5,
                      value=2,
                      command=old_form)
rb1_5_2.place(relx=0.65,
              rely=0.82)

b1_1 = Button(frame1,
              text='初始化',
              fg='#111111',
              width=6,
              command=para_init)
b1_1.place(relx=0.05,
           rely=0.92)

b1_2 = Button(frame1,
              text='生成',
              fg='#111111',
              width=6,
              command=lambda: vis_tab_it())
b1_2.place(relx=0.35,
           rely=0.92)

b1_3 = Button(frame1,
              text='RVR',
              fg='#111111',
              width=6,
              command=make_rvr_book)
b1_3.place(relx=0.65,
           rely=0.92)

frame2 = LabelFrame(root,
                    relief=SOLID)  # 框架
frame2.place(relx=0.31,
             rely=0.02,
             relheight=0.67,
             relwidth=0.68)

label2_1 = Label(frame2,
                 text='数据校对',
                 bg='#eeeeee',
                 fg='#111111',
                 justify=CENTER,
                 relief=FLAT)
label2_1.place(relx=0.01,
               rely=0.02,
               relwidth=0.98)

text_2 = Text(frame2,
              state=DISABLED)
text_2.place(relx=0.01,
             rely=0.07,
             relheight=0.80,
             relwidth=0.98)
update_txt(text_2,
           '此处显示需要校对的数据：')
sb_2 = Scrollbar(text_2,
                 command=text_2.yview)
sb_2.pack(side='right',
          fill='y')
text_2.config(yscrollcommand=sb_2.set)

frame2_1 = LabelFrame(frame2,
                      relief=GROOVE)

label2_1_1 = Label(frame2_1,
                   text='逐个校对',
                   bg='#eeeeee',
                   fg='#111111',
                   justify=CENTER,
                   relief=FLAT)
label2_1_1.place(relx=0.01,
                 rely=0.02,
                 relwidth=0.98)

label2_1_2 = Label(frame2_1,
                   text='年月/时次：',
                   bg='#eeeeee',
                   fg='#111111',
                   justify=LEFT,
                   relief=FLAT)
label2_1_2.place(relx=0.1,
                 rely=0.2)

label2_1_3 = Label(frame2_1,
                   text='当前取值：',
                   bg='#eeeeee',
                   fg='#111111',
                   justify=LEFT,
                   relief=FLAT)
label2_1_3.place(relx=0.1,
                 rely=0.4)

label2_1_4 = Label(frame2_1,
                   text='实际数据：',
                   bg='#eeeeee',
                   fg='#111111',
                   justify=LEFT,
                   relief=FLAT)
label2_1_4.place(relx=0.1,
                 rely=0.6)

v_2_1_5 = StringVar()
label2_1_5 = Label(frame2_1,
                   textvariable=v_2_1_5,
                   bg='#eeeeee',
                   fg='#111111',
                   justify=LEFT,
                   relief=FLAT)
label2_1_5.place(relx=0.45,
                 rely=0.2)

v_2_1_6 = StringVar()
label2_1_6 = Label(frame2_1,
                   textvariable=v_2_1_6,
                   bg='#eeeeee',
                   fg='#111111',
                   justify=LEFT,
                   relief=FLAT)
label2_1_6.place(relx=0.45,
                 rely=0.4)

entry2_1_1 = Entry(frame2_1)
entry2_1_1.place(relx=0.45,
                 rely=0.6,
                 relheight=0.07,
                 relwidth=0.3)

button2_1_1 = Button(frame2_1,
                     text='上一个',
                     fg='#123456',
                     command=goto_before)
button2_1_1.place(relx=0.1,
                  rely=0.8,
                  relwidth=0.2)

button2_1_2 = Button(frame2_1,
                     text='更新能见度',
                     fg='#123456',
                     command=update_vis)
button2_1_2.place(relx=0.4,
                  rely=0.8,
                  relwidth=0.2)

button2_1_3 = Button(frame2_1,
                     text='下一个',
                     fg='#123456',
                     command=goto_next)
button2_1_3.place(relx=0.7,
                  rely=0.8,
                  relwidth=0.2)

b0 = Button(frame2,
            text='存文本',
            fg='#123456',
            command=save_txt)
b0.place(relx=0.1,
         rely=0.9)

b2 = Button(frame2,
            text='清屏',
            fg='#123456',
            command=delete_text_2)
b2.place(relx=0.3,
         rely=0.9)

b3 = Button(frame2,
            text='开始校对',
            fg='#123456',
            command=start_update_vis)
b3.place(relx=0.5,
         rely=0.9)

b4 = Button(frame2,
            text='保存校对后的Excel',
            fg='#123456',
            command=save_change_excel)
b4.place(relx=0.7,
         rely=0.9)

root.mainloop()
#################################################################################################################
